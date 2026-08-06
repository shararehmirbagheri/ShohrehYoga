from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import arabic_reshaper
from bidi.algorithm import get_display

from kivy.clock import Clock
from kivy.graphics import Color, Line, Rectangle
from kivy.properties import (
    BooleanProperty,
    NumericProperty,
    ObjectProperty,
    StringProperty,
)
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.label import Label
from kivy.uix.screenmanager import Screen

from database import (
    Database,
    CLASS_HELD,
    CANCELED_CLASS,
    EXTRA_CLASS,
    NO_CLASS,
)
BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"

from jalali_calendar import (
    gregorian_iso_to_jalali_display,
    open_jalali_date_picker,
)


def display_persian(text):
    if not text:
        return ""
    return get_display(arabic_reshaper.reshape(str(text)))


class AttendanceStudentRow(ButtonBehavior, BoxLayout):
    student_id = NumericProperty(0)
    student_name = StringProperty("")
    present = BooleanProperty(False)
    class_status = StringProperty(CLASS_HELD)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "horizontal"
        self.size_hint_y = None
        self.height = 46
        self.spacing = 0

        with self.canvas.before:
            self._bg_color = Color(0.98, 0.97, 1, 1)
            self._bg_rect = Rectangle(pos=self.pos, size=self.size)
        with self.canvas.after:
            Color(0.83, 0.79, 0.88, 1)
            self._border = Line(points=[], width=0.7)

        self.checkbox = CheckBox(
            active=self.present,
            size_hint_x=0.12,
        )
        self.checkbox.bind(active=self._checkbox_changed)
        self.add_widget(self.checkbox)

        name_label = Label(
            text=display_persian(self.student_name),
            font_name="Persian",
            font_size=15,
            color=(0.18, 0.10, 0.25, 1),
            size_hint_x=0.63,
            halign="right",
            valign="middle",
            padding=(10, 0),
        )
        name_label.bind(
            size=lambda widget, size: setattr(widget, "text_size", size)
        )
        self.add_widget(name_label)

        self.status_label = Label(
            text="Absent",
            font_name="Persian",
            font_size=14,
            color=(0.18, 0.10, 0.25, 1),
            size_hint_x=0.25,
            halign="center",
            valign="middle",
        )
        self.status_label.bind(
            size=lambda widget, size: setattr(widget, "text_size", size)
        )
        self.add_widget(self.status_label)

        self.bind(
            pos=self._update_canvas,
            size=self._update_canvas,
            present=self._refresh_state,
            class_status=self._refresh_state,
        )
        self._refresh_state()

    def _update_canvas(self, *_):
        self._bg_rect.pos = self.pos
        self._bg_rect.size = self.size
        self._border.points = [
            self.x, self.y, self.right, self.y, self.right,
            self.top, self.x, self.top, self.x, self.y
        ]

    def _checkbox_changed(self, _, active):
        self.present = active

    def _refresh_state(self, *_):
        locked = self.class_status in {CANCELED_CLASS, NO_CLASS}
        self.checkbox.disabled = locked

        if locked:
            if self.checkbox.active:
                self.checkbox.active = False
            self.present = False
            self._bg_color.rgba = (0.85, 0.85, 0.87, 1)
            self.status_label.text = self.class_status
        elif self.present:
            self._bg_color.rgba = (0.86, 0.96, 0.87, 1)
            self.status_label.text = "Present"
        else:
            self._bg_color.rgba = (0.98, 0.97, 1, 1)
            self.status_label.text = "Absent"

    def on_release(self):
        if self.class_status not in {CANCELED_CLASS, NO_CLASS}:
            self.present = not self.present


class AttendanceScreen(Screen):
    attendance_list = ObjectProperty(None)
    selected_date_input = ObjectProperty(None)
    class_status_input = ObjectProperty(None)
    present_count_label = ObjectProperty(None)
    absent_count_label = ObjectProperty(None)
    total_count_label = ObjectProperty(None)
    message_label = ObjectProperty(None)

    sort_column = StringProperty("name")
    sort_reverse = BooleanProperty(False)

    def on_enter(self):
        Clock.schedule_once(self._initial_load, 0)

    def _initial_load(self, *_):
        if not self.selected_date_input or not self.attendance_list:
            return

        today_iso = date.today().isoformat()
        if not getattr(self.selected_date_input, "iso_date", ""):
            self.selected_date_input.iso_date = today_iso
            self.selected_date_input.text = (
                gregorian_iso_to_jalali_display(today_iso)
            )
        self.load_attendance()

    def _selected_iso(self):
        value = getattr(self.selected_date_input, "iso_date", "")
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return value
        except (TypeError, ValueError):
            value = date.today().isoformat()
            self.selected_date_input.iso_date = value
            self.selected_date_input.text = (
                gregorian_iso_to_jalali_display(value)
            )
            return value

    def _set_message(self, text):
        if self.message_label is not None:
            self.message_label.text = text

    def _current_class_status(self):
        if self.class_status_input is not None:
            value = str(self.class_status_input.text or "").strip()
            if value:
                return value
        return Database().get_class_status(self._selected_iso())

    def open_calendar(self):
        current = self._selected_iso()

        def apply_date(gregorian_iso, jalali_display):
            self.selected_date_input.iso_date = gregorian_iso
            self.selected_date_input.text = jalali_display
            self.load_attendance()

        open_jalali_date_picker(
            current_gregorian_iso=current,
            on_select=apply_date,
            title="انتخاب تاریخ حضور",
        )

    def use_today(self):
        today_iso = date.today().isoformat()
        self.selected_date_input.iso_date = today_iso
        self.selected_date_input.text = (
            gregorian_iso_to_jalali_display(today_iso)
        )
        self.load_attendance()

    def load_attendance(self):
        attended_on = self._selected_iso()
        db = Database()
        class_status = db.get_class_status(attended_on)

        if self.class_status_input:
            self.class_status_input.text = class_status

        present_ids = db.get_attendance_for_date(attended_on)
        students = db.get_students()

        records = [
            {
                "student_id": student_id,
                "name": name,
                "present": student_id in present_ids,
            }
            for student_id, name, phone, notes in students
        ]
        records.sort(key=self._sort_key, reverse=self.sort_reverse)

        self.attendance_list.clear_widgets()
        for record in records:
            row = AttendanceStudentRow(
                student_id=record["student_id"],
                student_name=record["name"],
                present=record["present"],
                class_status=class_status,
            )
            row.bind(present=lambda *_: self.update_counts())
            self.attendance_list.add_widget(row)

        self._set_message(
            "Thursday/Friday: no regular class."
            if class_status == NO_CLASS
            else "Class canceled; attendance will not reduce sessions."
            if class_status == CANCELED_CLASS
            else ""
        )
        self.update_counts()

    def on_class_status_changed(self, value):
        if not self.attendance_list:
            return
        for row in self.attendance_list.children:
            if isinstance(row, AttendanceStudentRow):
                row.class_status = value
        self.update_counts()

    def _sort_key(self, record):
        if self.sort_column == "present":
            return int(record["present"])
        if self.sort_column == "status":
            return "Present" if record["present"] else "Absent"
        return str(record["name"] or "").casefold()

    def sort_attendance(self, column):
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        self.load_attendance()

    def select_all(self):
        if self._current_class_status() in {CANCELED_CLASS, NO_CLASS}:
            return
        for row in self.attendance_list.children:
            if isinstance(row, AttendanceStudentRow):
                row.present = True
        self.update_counts()

    def clear_all(self):
        for row in self.attendance_list.children:
            if isinstance(row, AttendanceStudentRow):
                row.present = False
        self.update_counts()

    def save_attendance(self):
        attended_on = self._selected_iso()
        class_status = self._current_class_status()

        present_ids = [
            row.student_id
            for row in self.attendance_list.children
            if isinstance(row, AttendanceStudentRow) and row.present
        ]

        Database().save_attendance_for_date(
            attended_on,
            present_ids,
            class_status,
        )

        self._set_message(
            f"Saved {class_status} for "
            f"{self.selected_date_input.text}"
        )

        if self.manager and self.manager.has_screen("students"):
            self.manager.get_screen("students").load_students()

        self.load_attendance()

    def update_counts(self):
        rows = [
            row
            for row in self.attendance_list.children
            if isinstance(row, AttendanceStudentRow)
        ]
        present = sum(1 for row in rows if row.present)

        if self._current_class_status() in {CANCELED_CLASS, NO_CLASS}:
            absent = 0
            present = 0
        else:
            absent = len(rows) - present

        if self.total_count_label is not None:
            self.total_count_label.text = f"Students: {len(rows)}"
        if self.present_count_label is not None:
            self.present_count_label.text = f"Present: {present}"
        if self.absent_count_label is not None:
            self.absent_count_label.text = f"Absent: {absent}"

    def export_attendance(self):
        """
        Backward-compatible export for older KV files.
        Exports the currently selected date to Excel.
        """
        attended_on = self._selected_iso()
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        output = EXPORT_DIR / f"attendance_{attended_on}.xlsx"

        rows = Database().get_attendance_export_rows(attended_on)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(
            ["Date", "Student ID", "Name", "Phone", "Class Status", "Status"]
        )

        for row in rows:
            sheet.append(
                [
                    gregorian_iso_to_jalali_display(row["date"]),
                    row["student_id"],
                    row["name"],
                    str(row["phone"]),
                    row["class_status"],
                    row["status"],
                ]
            )

        header_fill = PatternFill(fill_type="solid", fgColor="6B28B8")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                if cell.column == 4:
                    cell.number_format = "@"
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = min(
                max(max_length + 2, 12),
                35,
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(output)
        self._set_message(f"Saved: {output.name}")

    def backup_database(self):
        output = Database().backup_database()
        self._set_message(f"Backup: {output.name}")
