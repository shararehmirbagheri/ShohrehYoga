from pathlib import Path

from openpyxl import load_workbook

from database import Database

EXCEL_FILE = Path("students.xlsx")
if not EXCEL_FILE.exists():
    EXCEL_FILE = Path("Students.xlsx")

WORKSHEET_NAME = "لیست حضور و غیاب"

PERSIAN_MONTHS = {
    "فروردین": 1,
    "اردیبهشت": 2,
    "خرداد": 3,
    "تیر": 4,
    "مرداد": 5,
    "شهریور": 6,
    "مهر": 7,
    "آبان": 8,
    "آذر": 9,
    "دی": 10,
    "بهمن": 11,
    "اسفند": 12,
}

DIGIT_MAP = str.maketrans(
    "۰۱۲۳۴۵۶۷۸۹",
    "0123456789",
)


def to_ascii_digits(value):
    if not isinstance(value, str):
        return value
    return value.translate(DIGIT_MAP)


def parse_month_name(value):
    if not value:
        return None
    text = str(value).strip()
    text = text.replace("\u200c", " ")
    text = text.replace("\n", " ")
    for month_name, month_index in PERSIAN_MONTHS.items():
        if month_name in text:
            return month_index
    return None


def parse_header_day(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u200c", " ")
    text = text.replace("\n", " ")
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def build_attendance_dates(ws, default_year=1405):
    dates = [None] * len(ws[1])
    current_month = None

    for idx, month_cell in enumerate(ws[1]):
        month_index = parse_month_name(month_cell.value)
        if month_index is not None:
            current_month = month_index

        if idx < len(ws[2]):
            day = parse_header_day(ws[2][idx].value)
            if current_month and day:
                dates[idx] = f"{default_year}/{current_month:02}/{day:02}"

    return dates


def row_has_attendance(cell_value):
    if cell_value is None:
        return False
    if isinstance(cell_value, str):
        return bool(cell_value.strip())
    if isinstance(cell_value, (int, float)):
        return True
    return True


if not EXCEL_FILE.exists():
    print("students.xlsx or Students.xlsx not found!")
    exit(1)

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[WORKSHEET_NAME] if WORKSHEET_NAME in wb.sheetnames else wb.active

print(f"Using worksheet: {ws.title}")

attendance_dates = build_attendance_dates(ws)
attendance_columns = [idx + 1 for idx, date_value in enumerate(attendance_dates) if date_value]
print(f"Found {len(attendance_columns)} attendance columns.")

if not attendance_columns:
    print("No attendance columns could be mapped from the worksheet headers.")
    exit(1)


db = Database()
student_map = {row[1].strip(): row[0] for row in db.get_students() if row[1]}

added_students = 0
skipped_students = 0
added_attendances = 0
skipped_attendances = 0

for row in ws.iter_rows(min_row=3, values_only=True):
    name = row[1] if len(row) > 1 else None
    if not name:
        continue

    name = str(name).strip()
    if not name:
        continue

    if name not in student_map:
        db.add_student(name)
        student_map[name] = db.cursor.lastrowid
        added_students += 1
    else:
        skipped_students += 1

    student_id = student_map[name]

    for idx, attendance_date in enumerate(attendance_dates):
        if not attendance_date:
            continue
        if idx >= len(row):
            continue

        if not row_has_attendance(row[idx]):
            continue

        if db.attendance_exists(student_id, attendance_date):
            skipped_attendances += 1
            continue

        db.add_attendance(student_id, attendance_date)
        added_attendances += 1

print("--------------------------------")
print(f"New students added       : {added_students}")
print(f"Existing students skipped: {skipped_students}")
print(f"Attendance records added : {added_attendances}")
print(f"Attendance records skipped: {skipped_attendances}")
print("--------------------------------")
print("Import Complete.")
