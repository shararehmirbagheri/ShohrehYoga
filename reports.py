from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from kivy.graphics import Color, Line, Rectangle
from kivy.properties import ObjectProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.screenmanager import Screen

from database import Database, CANCELED_CLASS
from jalali_calendar import (
    gregorian_iso_to_jalali_display,
    open_jalali_date_picker,
)

BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"

PERSIAN_WEEKDAYS = {
    0: "دوشنبه",
    1: "سه‌شنبه",
    2: "چهارشنبه",
    3: "پنجشنبه",
    4: "جمعه",
    5: "شنبه",
    6: "یکشنبه",
}


class ReportCell(Label):
    """A bordered preview-table cell."""

    def __init__(
        self,
        background=(1, 1, 1, 1),
        text_color=(0.16, 0.10, 0.23, 1),
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.color = text_color
        self.font_name = "Persian"
        self.font_size = 12
        self.halign = "center"
        self.valign = "middle"
        self.padding = (5, 3)
        self.text_size = self.size

        with self.canvas.before:
            self._background_color = Color(*background)
            self._background = Rectangle(pos=self.pos, size=self.size)

        with self.canvas.after:
            Color(0.73, 0.70, 0.78, 1)
            self._border = Line(rectangle=(self.x, self.y, self.width, self.height), width=0.8)

        self.bind(pos=self._update_canvas, size=self._update_canvas)

    def _update_canvas(self, *_):
        self._background.pos = self.pos
        self._background.size = self.size
        self._border.rectangle = (self.x, self.y, self.width, self.height)
        self.text_size = self.size


class ReportScreen(Screen):
    from_date_input = ObjectProperty(None)
    to_date_input = ObjectProperty(None)
    report_message = ObjectProperty(None)
    preview_grid = ObjectProperty(None)

    def on_enter(self):
        today = date.today()
        first_day = today.replace(day=1)

        if not getattr(self.from_date_input, "iso_date", ""):
            self.from_date_input.iso_date = first_day.isoformat()
            self.from_date_input.text = gregorian_iso_to_jalali_display(
                first_day.isoformat()
            )

        if not getattr(self.to_date_input, "iso_date", ""):
            self.to_date_input.iso_date = today.isoformat()
            self.to_date_input.text = gregorian_iso_to_jalali_display(
                today.isoformat()
            )

    def _range(self):
        start_value = getattr(self.from_date_input, "iso_date", "")
        end_value = getattr(self.to_date_input, "iso_date", "")

        try:
            start = datetime.strptime(start_value, "%Y-%m-%d").date()
            end = datetime.strptime(end_value, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            end = date.today()
            start = end.replace(day=1)

        if start > end:
            start, end = end, start

        self.from_date_input.iso_date = start.isoformat()
        self.from_date_input.text = gregorian_iso_to_jalali_display(
            start.isoformat()
        )
        self.to_date_input.iso_date = end.isoformat()
        self.to_date_input.text = gregorian_iso_to_jalali_display(
            end.isoformat()
        )

        return start.isoformat(), end.isoformat()

    def open_from_calendar(self):
        current = getattr(
            self.from_date_input,
            "iso_date",
            date.today().isoformat(),
        )

        def apply_date(gregorian_iso, jalali_display):
            self.from_date_input.iso_date = gregorian_iso
            self.from_date_input.text = jalali_display

        open_jalali_date_picker(
            current_gregorian_iso=current,
            on_select=apply_date,
            title="انتخاب تاریخ شروع گزارش",
        )

    def open_to_calendar(self):
        current = getattr(
            self.to_date_input,
            "iso_date",
            date.today().isoformat(),
        )

        def apply_date(gregorian_iso, jalali_display):
            self.to_date_input.iso_date = gregorian_iso
            self.to_date_input.text = jalali_display

        open_jalali_date_picker(
            current_gregorian_iso=current,
            on_select=apply_date,
            title="انتخاب تاریخ پایان گزارش",
        )

    @staticmethod
    def _cell_widths(report_dates):
        static = [58, 220, 82, 120, 90, 90, 95]
        dates = [105] * len(report_dates)
        return static + dates

    def view_report(self):
        start_iso, end_iso = self._range()
        report_dates, students = Database().get_attendance_matrix(
            start_iso,
            end_iso,
        )

        self.preview_grid.clear_widgets()

        widths = self._cell_widths(report_dates)
        self.preview_grid.cols = len(widths)
        self.preview_grid.width = sum(widths)
        self.preview_grid.row_default_height = 42
        self.preview_grid.height = (len(students) + 1) * 42

        headers = [
            "ID",
            "Student Name",
            "Receipt",
            "First Session",
            "Paid",
            "Used",
            "Remaining",
        ]

        for item in report_dates:
            iso_date = item["date"]
            gregorian = datetime.strptime(iso_date, "%Y-%m-%d").date()
            weekday = PERSIAN_WEEKDAYS[gregorian.weekday()]
            date_text = gregorian_iso_to_jalali_display(iso_date)
            header = f"{weekday}\n{date_text}"
            if item["status"] == CANCELED_CLASS:
                header += "\nCanceled"
            headers.append(header)

        for index, header in enumerate(headers):
            self.preview_grid.add_widget(
                ReportCell(
                    text=header,
                    size_hint=(None, None),
                    width=widths[index],
                    height=42,
                    bold=True,
                    background=(0.42, 0.16, 0.72, 1),
                    text_color=(1, 1, 1, 1),
                )
            )

        for student in students:
            if student["receipt"] == "No":
                row_background = (1.00, 0.95, 0.66, 1)
            elif student["remaining_sessions"] < 0:
                row_background = (1.00, 0.78, 0.78, 1)
            elif student["remaining_sessions"] <= 3:
                row_background = (1.00, 0.88, 0.65, 1)
            else:
                row_background = (1, 1, 1, 1)

            first_session = (
                gregorian_iso_to_jalali_display(student["first_session_date"])
                if student["first_session_date"]
                else "*"
            )

            static_values = [
                str(student["id"]),
                student["name"],
                "Yes" if student["receipt"] == "Yes" else "No",
                first_session,
                str(student["sessions_paid"]),
                str(student["sessions_used"]),
                str(student["remaining_sessions"]),
            ]

            for index, value in enumerate(static_values):
                self.preview_grid.add_widget(
                    ReportCell(
                        text=value,
                        size_hint=(None, None),
                        width=widths[index],
                        height=42,
                        background=row_background,
                    )
                )

            first_date_index = len(static_values)
            for date_offset, item in enumerate(report_dates):
                state = student["attendance"][item["date"]]

                if state == "Present":
                    value = "✓"
                    background = (0.57, 0.82, 0.45, 1)
                elif state == "Canceled":
                    value = "Canceled"
                    background = (0.76, 0.76, 0.78, 1)
                else:
                    value = ""
                    background = (1, 1, 1, 1)

                self.preview_grid.add_widget(
                    ReportCell(
                        text=value,
                        size_hint=(None, None),
                        width=widths[first_date_index + date_offset],
                        height=42,
                        background=background,
                    )
                )

        canceled = sum(
            1 for item in report_dates
            if item["status"] == CANCELED_CLASS
        )
        self.report_message.text = (
            f"Students: {len(students)} | "
            f"Class columns: {len(report_dates)} | "
            f"Canceled: {canceled}"
        )

    def export_report(self):
        start_iso, end_iso = self._range()
        report_dates, students = Database().get_attendance_matrix(
            start_iso,
            end_iso,
        )

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        output = EXPORT_DIR / (
            f"attendance_matrix_{start_iso}_to_{end_iso}.xlsx"
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance Matrix"
        sheet.sheet_view.rightToLeft = True

        static_headers = [
            "ردیف",
            "نام و نام خانوادگی",
            "فیش",
            "تاریخ جلسه اول",
            "جلسات خریداری",
            "جلسات محاسبه شده",
            "جلسات مانده",
            "تلفن",
        ]

        for column, header in enumerate(static_headers, start=1):
            sheet.cell(row=1, column=column, value=header)

        first_date_column = len(static_headers) + 1

        for offset, item in enumerate(report_dates):
            column = first_date_column + offset
            iso_date = item["date"]
            gregorian = datetime.strptime(iso_date, "%Y-%m-%d").date()
            jalali_date = gregorian_iso_to_jalali_display(iso_date)
            weekday = PERSIAN_WEEKDAYS[gregorian.weekday()]

            value = f"{weekday}\n{jalali_date}"
            if item["status"] == CANCELED_CLASS:
                value += "\nلغو"

            sheet.cell(row=1, column=column, value=value)

        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        violet = PatternFill("solid", fgColor="6B28B8")
        green = PatternFill("solid", fgColor="92D050")
        gray = PatternFill("solid", fgColor="BFBFBF")
        yellow = PatternFill("solid", fgColor="FFF2CC")
        orange = PatternFill("solid", fgColor="F4B183")
        red = PatternFill("solid", fgColor="F4CCCC")
        regular_font = Font(name="Tahoma", size=10)
        header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")

        for cell in sheet[1]:
            cell.fill = violet
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        for row_index, student in enumerate(students, start=2):
            first_session = (
                gregorian_iso_to_jalali_display(student["first_session_date"])
                if student["first_session_date"]
                else "*"
            )

            static_values = [
                row_index - 1,
                student["name"],
                "دارد" if student["receipt"] == "Yes" else "ندارد",
                first_session,
                student["sessions_paid"],
                student["sessions_used"],
                student["remaining_sessions"],
                str(student["phone"] or ""),
            ]

            if student["receipt"] == "No":
                row_fill = yellow
            elif student["remaining_sessions"] < 0:
                row_fill = red
            elif student["remaining_sessions"] <= 3:
                row_fill = orange
            else:
                row_fill = None

            for column, value in enumerate(static_values, start=1):
                cell = sheet.cell(row=row_index, column=column, value=value)
                cell.font = regular_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border
                if row_fill:
                    cell.fill = row_fill

            sheet.cell(row=row_index, column=2).alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
            sheet.cell(row=row_index, column=8).number_format = "@"

            for offset, item in enumerate(report_dates):
                column = first_date_column + offset
                state = student["attendance"][item["date"]]
                cell = sheet.cell(row=row_index, column=column)
                cell.font = regular_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                cell.border = border

                if state == "Present":
                    cell.value = "✓"
                    cell.fill = green
                elif state == "Canceled":
                    cell.value = "لغو"
                    cell.fill = gray
                else:
                    cell.value = ""

        sheet.freeze_panes = "I2"
        sheet.row_dimensions[1].height = 48
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 16
        sheet.column_dimensions["E"].width = 14
        sheet.column_dimensions["F"].width = 16
        sheet.column_dimensions["G"].width = 13
        sheet.column_dimensions["H"].width = 16

        for column in range(first_date_column, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 12

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:1"
        sheet.auto_filter.ref = sheet.dimensions

        workbook.save(output)
        self.report_message.text = f"Saved: {output.name}"
